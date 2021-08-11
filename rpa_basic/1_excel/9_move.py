from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
wb = load_workbook('sample.xlsx')
ws: Worksheet = wb.active

# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value = "국어"

# ws.move_range("C1:C11", rows=5, cols=-1)

wb.save("sample_korean.xlsx")
