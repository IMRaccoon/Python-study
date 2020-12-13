from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
wb = load_workbook('sample.xlsx')
ws: Worksheet = wb.active


# 번호, 영어, 수학
a1 = ws["A1"]   # 번호
b1 = ws["B1"]   # 영어
c1 = ws["C1"]   # 수학

# A 열의 너비를 5로 설정
ws.column_dimensions['A'].width = 5
# 1 행의 높이를 50 으로 설정
ws.row_dimensions[1].height = 50

# 스타일 적용
# 글자 색은 빨갛게, 이탤릭, 두껩게 적용
a1.font = Font(color="FF0000", italic=True, bold=True)
# 글자 색은 보라색, 폰트 Arial, 취소선
b1.font = Font(color="CC33FF", name="Arial", strike=True)
# 글자 색은 파란색, 글씨 크기는 20, 밑줄선
c1.font = Font(color="0000FF", size=20, underline='single')


wb.save("sample_style.xlsx")
