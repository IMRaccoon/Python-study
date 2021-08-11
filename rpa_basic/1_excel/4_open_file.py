from openpyxl import load_workbook  # 파일 불러오기
wb = load_workbook('sample.xlsx')   # sample.xlsx 파일엣 wb를 불러옴
ws = wb.active

# cell 데이터 불러오기
# for x in range(1, 11):
#     for y in range(1, 11):
#         print(ws.cell(row=x, column=y).value, end=" ")  # 1, 2, 3, ...
#     print()

# cell 갯수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")  # 1, 2, 3, ...
    print()